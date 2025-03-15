import os
import re
import json
import docx
import PyPDF2
import openai
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Renovation Transcript to Excel",
    page_icon="📝",
    layout="wide",
)

st.markdown("""
<style>
.block-container {
    max-width: 900px;
}
h1, h2, h3, h4 {
    font-family: "Segoe UI", Tahoma, Geneva, Verdana, sans-serif;
}
</style>
""", unsafe_allow_html=True)

openai.api_key = os.getenv("OPENAI_API_KEY")
if not openai.api_key:
    st.warning("Please set your OPENAI_API_KEY as an environment variable.")
GPT_MODEL = "gpt-4"

def extract_text(file_bytes, file_extension):
    file_extension = file_extension.lower()
    if file_extension == ".docx":
        return extract_text_from_docx(file_bytes)
    elif file_extension == ".pdf":
        return extract_text_from_pdf(file_bytes)
    else:
        raise ValueError("Unsupported file format: " + file_extension)

def extract_text_from_docx(file_bytes):
    document = docx.Document(file_bytes)
    paragraphs = [para.text for para in document.paragraphs]
    return "\n".join(paragraphs)

def extract_text_from_pdf(file_bytes):
    text = ""
    pdf_reader = PyPDF2.PdfReader(file_bytes)
    for page in pdf_reader.pages:
        page_text = page.extract_text() or ""
        text += page_text + "\n"
    return text

def clean_text(text):
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def build_prompt(transcript_text):
    prompt = f"""
You are an AI assistant extracting renovation details from a client transcript.
Please carefully analyze the conversation and return a JSON object with these keys:

1. "ProjectName": (If not mentioned, return "Not provided")
2. "ClientName": (If not mentioned, return "Not provided")
3. "PropertyAddress": (If not mentioned, return "Not provided")
4. "ProjectManager": (If not mentioned, return "Not provided")
5. "RenovationAreas": (List or describe the rooms/areas, e.g., "Kitchen, Bathroom")
6. "ScopeOfWork": (Summarize all renovation tasks or goals)
7. "MaterialPreferences": (List any specific materials or design preferences)
8. "BudgetOrCost": (Any budget or cost references)
9. "Timeline": (Any schedule or start/end dates mentioned)
10. "AdditionalNotes": (Extra details like permit requirements, constraints, etc.)

Transcript:
{transcript_text}

Return only valid JSON with exactly the keys:
ProjectName, ClientName, PropertyAddress, ProjectManager, RenovationAreas, ScopeOfWork, MaterialPreferences, BudgetOrCost, Timeline, AdditionalNotes.
    """
    return prompt

def extract_details_with_gpt(transcript_text):
    prompt = build_prompt(transcript_text)
    response = openai.ChatCompletion.create(
        model=GPT_MODEL,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2
    )
    gpt_output = response.choices[0].message.content.strip()
    try:
        details = json.loads(gpt_output)
    except json.JSONDecodeError as e:
        st.error("Error parsing GPT JSON output. Returning fallback data.")
        st.write("Raw GPT output was:", gpt_output)
        details = {
            "ProjectName": "Not provided",
            "ClientName": "Not provided",
            "PropertyAddress": "Not provided",
            "ProjectManager": "Not provided",
            "RenovationAreas": "Not provided",
            "ScopeOfWork": "Not provided",
            "MaterialPreferences": "Not provided",
            "BudgetOrCost": "Not provided",
            "Timeline": "Not provided",
            "AdditionalNotes": "Not provided"
        }
    return details

def create_pretty_excel(details):
    wb = Workbook()
    ws = wb.active
    ws.title = "Renovation Data"
    headers = [
        "ProjectName", "ClientName", "PropertyAddress", "ProjectManager",
        "RenovationAreas", "ScopeOfWork", "MaterialPreferences",
        "BudgetOrCost", "Timeline", "AdditionalNotes"
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 25
    for col_idx, header in enumerate(headers, start=1):
        val = details.get(header, "Not provided")
        if isinstance(val, list):
            val = ", ".join(val)
        cell = ws.cell(row=2, column=col_idx, value=str(val))
        cell.alignment = wrap_alignment
        cell.border = thin_border
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 25
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def main():
    st.sidebar.title("Chapter Renovation Details")
    st.sidebar.markdown("### Upload your transcript here:")
    uploaded_file = st.sidebar.file_uploader("Choose a .docx or .pdf file", type=["docx", "pdf"])
    st.title("Renovation Transcript ↦ Excel Generator")
    st.write(
        "This app takes a DOCX or PDF transcript about a renovation project, "
        "extracts the key details using GPT, and returns a downloadable Excel file."
    )
    if uploaded_file is not None:
        file_extension = os.path.splitext(uploaded_file.name)[1]
        st.write(f"**File uploaded:** `{uploaded_file.name}`")
        process_button = st.button("Process File with GPT")
        if process_button:
            try:
                with st.spinner("Extracting text..."):
                    raw_text = extract_text(uploaded_file, file_extension)
                    cleaned_text = clean_text(raw_text)
                with st.spinner("Extracting details via GPT..."):
                    details = extract_details_with_gpt(cleaned_text)
                st.success("Extraction complete! Here are the details:")
                st.json(details)
                with st.spinner("Generating Excel file..."):
                    excel_data = create_pretty_excel(details)
                st.success("Excel file ready!")
                st.download_button(
                    label="Download Excel",
                    data=excel_data,
                    file_name="Renovation_Extracted_Details.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"An error occurred: {e}")
    else:
        st.info("Please upload a DOCX or PDF file from the sidebar to begin.")

if __name__ == "__main__":
    main()
